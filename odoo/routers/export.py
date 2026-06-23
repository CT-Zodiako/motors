import csv
import io
import json
import tempfile
from typing import Literal

import openpyxl
from fastapi import APIRouter, Body, HTTPException, Query
from fastapi.responses import StreamingResponse, FileResponse

from db import query as pg_query
from odoo_client import execute as odoo_execute

router = APIRouter(prefix="/export", tags=["export"])


def _fetch_data(name: str) -> tuple[dict, list[dict]]:
    rows = pg_query(
        "SELECT * FROM odoo_queries WHERE name = %s AND active = TRUE", (name,)
    )
    if not rows:
        raise HTTPException(status_code=404, detail=f"Query '{name}' not found or inactive")

    registered = rows[0]
    data = odoo_execute(
        registered["model"],
        registered["method"],
        [registered["domain"]],
        {"fields": registered["fields"], "limit": False},
    )
    return registered, data


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _infer_pg_type(value: object) -> str:
    if isinstance(value, bool):
        return "BOOLEAN"
    if isinstance(value, int):
        return "INTEGER"
    if isinstance(value, float):
        return "NUMERIC"
    if isinstance(value, (list, dict)):
        return "TEXT"
    return "TEXT"


def _infer_oracle_type(value: object) -> str:
    if isinstance(value, bool):
        return "NUMBER(1)"
    if isinstance(value, int):
        return "NUMBER"
    if isinstance(value, float):
        return "NUMBER"
    if isinstance(value, (list, dict)):
        return "CLOB"
    return "VARCHAR2(4000)"


def _build_create_table(table: str, columns: list[str], data: list[dict], target: str) -> list[str]:
    # Scan first non-None value per column to infer type
    col_types: dict[str, str] = {}
    for col in columns:
        for row in data:
            val = row.get(col)
            if val is not None:
                col_types[col] = (
                    _infer_pg_type(val) if target == "postgres" else _infer_oracle_type(val)
                )
                break
        else:
            col_types[col] = "TEXT" if target == "postgres" else "VARCHAR2(4000)"

    col_defs = ",\n".join(f"    {col}  {col_types[col]}" for col in columns)

    if target == "postgres":
        return [
            f"CREATE TABLE IF NOT EXISTS {table} (\n{col_defs}\n);",
            "",
        ]
    else:
        # Oracle < 23c doesn't support IF NOT EXISTS — wrap in anonymous block
        return [
            "BEGIN",
            f"  EXECUTE IMMEDIATE 'CREATE TABLE {table} (",
            *[f"    {col}  {col_types[col]}{',' if i < len(columns) - 1 else ''}" for i, col in enumerate(columns)],
            "  )';",
            "EXCEPTION",
            "  WHEN OTHERS THEN",
            "    IF SQLCODE = -955 THEN NULL; END IF; -- ORA-00955: table already exists",
            "END;",
            "/",
            "",
        ]


def _sql_value(value: object, target: str) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        if target == "oracle":
            return "1" if value else "0"
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    raw = str(value) if not isinstance(value, (list, dict)) else json.dumps(value, ensure_ascii=False)
    escaped = raw.replace("'", "''")
    return f"'{escaped}'"


# ─── CSV ──────────────────────────────────────────────────────────────────────

def _resolve_columns(all_columns: list[str], param: str | None) -> list[str]:
    if not param:
        return all_columns
    requested = [c.strip() for c in param.split(",") if c.strip()]
    return [c for c in requested if c in all_columns] or all_columns


# ─── CSV ──────────────────────────────────────────────────────────────────────

@router.get("/csv/{name}")
def export_csv(name: str, columns: str | None = Query(None)):
    _, data = _fetch_data(name)
    if not data:
        raise HTTPException(status_code=204, detail="No data to export")

    cols = _resolve_columns(list(data[0].keys()), columns)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for row in data:
        writer.writerow({col: _cell(row.get(col)) for col in cols})

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={name}.csv"},
    )


# ─── EXCEL ────────────────────────────────────────────────────────────────────

@router.get("/excel/{name}")
def export_excel(name: str, columns: str | None = Query(None)):
    registered, data = _fetch_data(name)
    if not data:
        raise HTTPException(status_code=204, detail="No data to export")

    cols = _resolve_columns(list(data[0].keys()), columns)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="1C1C1E")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    for col_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(data, 2):
        for col_idx, col in enumerate(cols, 1):
            val = row.get(col)
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx, col in enumerate(cols, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(col) + 4, 14)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{name}.xlsx",
    )


# ─── SQL ──────────────────────────────────────────────────────────────────────

@router.get("/sql/{name}")
def export_sql(
    name: str,
    target: Literal["postgres", "oracle"] = Query("postgres"),
    columns: str | None = Query(None),
):
    registered, data = _fetch_data(name)
    if not data:
        raise HTTPException(status_code=204, detail="No data to export")

    table = name
    cols = _resolve_columns(list(data[0].keys()), columns)
    col_list = ", ".join(cols)

    lines: list[str] = []
    lines.append(f"-- Generated for {target.upper()} | query: {name} | model: {registered['model']}")
    lines.append(f"-- {len(data)} records | columns: {col_list}\n")

    lines.extend(_build_create_table(table, cols, data, target))

    if target == "oracle":
        lines.append("BEGIN")
        for row in data:
            vals = ", ".join(_sql_value(row.get(col), "oracle") for col in cols)
            lines.append(f"  INSERT INTO {table} ({col_list}) VALUES ({vals});")
        lines.append("END;")
        lines.append("/")
    else:
        for row in data:
            vals = ", ".join(_sql_value(row.get(col), "postgres") for col in cols)
            lines.append(f"INSERT INTO {table} ({col_list}) VALUES ({vals});")

    content = "\n".join(lines)
    return StreamingResponse(
        iter([content]),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={name}_{target}.sql"},
    )


@router.post("/sql-preview")
def sql_preview(payload: dict = Body(...)) -> dict:
    table = payload.get("table")
    columns = payload.get("columns")
    rows = payload.get("rows", [])

    if not isinstance(table, str) or not table.strip():
        raise HTTPException(status_code=400, detail="table must be a non-empty string")
    if not isinstance(columns, list) or len(columns) == 0:
        raise HTTPException(status_code=400, detail="columns must be a non-empty list")
    if not all(isinstance(c, str) and c.strip() for c in columns):
        raise HTTPException(status_code=400, detail="each column must be a non-empty string")
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="rows must be a list")

    target = "postgres"
    col_list = ", ".join(columns)

    lines: list[str] = []
    lines.extend(_build_create_table(table, columns, rows, target))

    for row in rows:
        vals = ", ".join(_sql_value(row.get(col), target) for col in columns)
        lines.append(f"INSERT INTO {table} ({col_list}) VALUES ({vals});")

    return {"sql": "\n".join(lines)}
