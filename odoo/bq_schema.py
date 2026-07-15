"""Pure domain logic for file-upload ingestion: extraction, sanitization, inference, conversion.

No FastAPI/HTTP imports. Reuses private string-inference helpers from routers.bigquery.
"""

import csv
import io
import math
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timezone

from routers.bigquery import (
    MAX_UPLOAD_ROWS,
    _infer_string_type,
)

CSV_SHEET_NAME = "CSV"

# BigQuery type ranking for promotion (higher = more permissive).
# Must stay in sync with _BQ_TYPE_RANK from routers.bigquery.
_BQ_TYPE_RANK = {
    "BOOLEAN": 0,
    "INTEGER": 1,
    "FLOAT": 2,
    "NUMERIC": 3,
    "BIGNUMERIC": 4,
    "TIMESTAMP": 5,
    "DATE": 6,
    "TIME": 7,
    "DATETIME": 8,
    "STRING": 9,
}

# Mapping from our internal type names to BigQuery type names.
# Our internal names use INT64/FLOAT64/BOOL; bigquery.py uses INTEGER/FLOAT/BOOLEAN.
_BQ_TYPE_MAP = {
    "INT64": "INTEGER",
    "FLOAT64": "FLOAT",
    "BOOL": "BOOLEAN",
    "DATE": "DATE",
    "TIMESTAMP": "TIMESTAMP",
    "STRING": "STRING",
    # Also map bigquery.py names to themselves for reverse lookup
    "INTEGER": "INTEGER",
    "FLOAT": "FLOAT",
    "BOOLEAN": "BOOLEAN",
}

# Reverse mapping for converting from bigquery.py types back to our types.
_BQ_TYPE_REVERSE = {
    "INTEGER": "INT64",
    "FLOAT": "FLOAT64",
    "BOOLEAN": "BOOL",
    "DATE": "DATE",
    "TIMESTAMP": "TIMESTAMP",
    "STRING": "STRING",
}


@dataclass
class ExtractedTable:
    """Unified extraction result for both xlsx and csv."""

    headers: list[str]
    rows: list[list]
    sheet_name: str = CSV_SHEET_NAME


class ConversionError(Exception):
    """Raised when a value cannot be converted to the target BigQuery type."""

    def __init__(self, column: str, row: int, value, target: str, reason: str = ""):
        self.column = column
        self.row = row
        self.value = value
        self.target = target
        self.reason = reason or f"value {value!r} is not compatible with {target}"
        super().__init__(f"Column '{column}' row {row}: {self.reason}")


class ExtractionError(Exception):
    """Raised when a file cannot be extracted (corrupt, ragged, empty, etc.)."""

    pass


# ── Sanitization ────────────────────────────────────────────────────


def sanitize_column_name(raw: str, used: set[str]) -> str:
    """Sanitize a raw column name to a valid BigQuery identifier.

    Rules (D8):
    - Replace every character outside [A-Za-z0-9_] with "_".
    - If empty or first character is not a letter or "_", prefix "_".
    - Truncate to 1024 characters.
    - Deduplicate case-insensitively by appending "_2", "_3", etc.
      Re-truncate the base so total length ≤ 1024.
    """
    if raw == "":
        candidate = "_"
    else:
        candidate = re.sub(r"[^A-Za-z0-9_]", "_", raw)
        if not re.match(r"^[A-Za-z_]", candidate):
            candidate = "_" + candidate
        candidate = candidate[:1024]

    # Deduplicate case-insensitively
    lower = candidate.lower()
    if lower not in used:
        used.add(lower)
        return candidate

    suffix = 2
    while True:
        suffix_str = f"_{suffix}"
        # Make room for suffix by truncating base
        base = candidate[: 1024 - len(suffix_str)]
        candidate_with_suffix = base + suffix_str
        lower_with_suffix = candidate_with_suffix.lower()
        if lower_with_suffix not in used:
            used.add(lower_with_suffix)
            return candidate_with_suffix
        suffix += 1


# ── CSV extraction ──────────────────────────────────────────────────


def _validate_skip_rows(skip_rows: int) -> None:
    if skip_rows < 0:
        raise ExtractionError(f"skip_rows must be >= 0, got {skip_rows}")


def extract_csv(content: bytes, skip_rows: int = 0) -> ExtractedTable:
    """Extract a CSV file into an ExtractedTable (D4).

    - Decode utf-8-sig (handles BOM); fall back to cp1252 on UnicodeDecodeError.
    - Sniff delimiter from {, ; \t |}; fallback comma.
    - Drop fully-empty rows.
    - Ragged rows (field count != header count) → ExtractionError with 1-based row number.
    - Empty file (no records) → ExtractionError.
    """
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1252", errors="replace")

    # Sniff delimiter from a sample
    sample = text[:8192] if len(text) > 8192 else text
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # fallback to comma

    reader = csv.reader(io.StringIO(text, newline=""), dialect=dialect)
    records = list(reader)

    if not records:
        raise ExtractionError("CSV file is empty or has no readable records")

    _validate_skip_rows(skip_rows)
    if skip_rows:
        if skip_rows >= len(records):
            raise ExtractionError(
                f"skip_rows={skip_rows} consumes all rows of the file"
            )
        records = records[skip_rows:]

    headers = records[0]
    expected = len(headers)
    rows = []
    for idx, record in enumerate(records[1:], start=1):
        # Drop fully-empty rows
        if all(cell == "" for cell in record):
            continue
        if len(record) != expected:
            raise ExtractionError(
                f"CSV row {idx} has {len(record)} fields, expected {expected}"
            )
        # Convert empty strings to None
        rows.append([cell if cell != "" else None for cell in record])

    return ExtractedTable(headers=headers, rows=rows, sheet_name=CSV_SHEET_NAME)


# ── xlsx extraction ─────────────────────────────────────────────────


def extract_xlsx(
    content: bytes, sheet_name: str | None, skip_rows: int = 0
) -> ExtractedTable | list[str]:
    """Extract an .xlsx file.

    If sheet_name is None, return the list of sheet names (for inspect).
    Otherwise, return an ExtractedTable for the specified sheet (D5).

    - load_workbook(read_only=True, data_only=True).
    - Sheet order from wb.sheetnames.
    - DATE vs TIMESTAMP via number_format time tokens [hs] after stripping
      quoted literals and [...] sections; midnight fallback.
    - Formula without cached value → None.
    - Short rows padded with None; long rows with non-empty extras → error.
    - Corrupt file → ExtractionError.
    - datetime.time cells → STRING isoformat.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ExtractionError(f"corrupt or unreadable .xlsx: {exc}") from exc

    if sheet_name is None:
        return wb.sheetnames

    if sheet_name not in wb.sheetnames:
        raise ExtractionError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    # Reset dimensions in case cached values are stale
    try:
        ws.reset_dimensions()
    except Exception:
        pass

    _validate_skip_rows(skip_rows)
    headers = None
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx <= skip_rows:
            continue
        values = []
        for cell in row:
            val = _extract_xlsx_cell_value(cell)
            values.append(val)

        if headers is None:
            headers = values
            continue

        # Drop fully-empty rows
        if all(v is None for v in values):
            continue

        # Pad short rows
        if len(values) < len(headers):
            values.extend([None] * (len(headers) - len(values)))

        # Check long rows: drop trailing empty cells (formatting artifacts)
        if len(values) > len(headers):
            extras = values[len(headers) :]
            if all(v is None for v in extras):
                values = values[: len(headers)]
            else:
                raise ExtractionError(
                    f"Row {row_idx - 1} has {len(values)} fields, expected {len(headers)}"
                )

        rows.append(values)

    if headers is None:
        raise ExtractionError(
            f"skip_rows={skip_rows} consumes all rows of sheet '{sheet_name}'"
        )

    return ExtractedTable(headers=headers, rows=rows, sheet_name=sheet_name)


def _extract_xlsx_cell_value(cell):
    """Extract the native Python value from an openpyxl cell (D5)."""
    value = cell.value
    if value is None:
        return None

    # datetime.time → STRING isoformat
    if isinstance(value, time):
        return value.isoformat()

    # datetime.datetime — decide DATE vs TIMESTAMP based on number_format
    if isinstance(value, datetime):
        fmt = cell.number_format or ""
        if _is_date_only_format(fmt):
            return value.date()
        return value

    # bool — must be checked before int (isinstance(True, int) is True)
    if isinstance(value, bool):
        return value

    # int, float, str pass through
    return value


def _is_date_only_format(fmt: str) -> bool:
    """Return True if the Excel number format represents a date without time (D5).

    Strips quoted literals and [...] sections, then checks for time tokens [hs].
    If no format is available or ambiguous, midnight → DATE fallback.
    """
    if not fmt:
        return True  # midnight fallback when no format

    # Strip quoted literals: "text" or 'text'
    cleaned = re.sub(r'"[^"]*"', "", fmt)
    cleaned = re.sub(r"'[^']*'", "", cleaned)
    # Strip [...] sections (locale/color codes)
    cleaned = re.sub(r"\[[^\]]*\]", "", cleaned)

    # If time tokens present, it's a TIMESTAMP candidate
    if re.search(r"[hs]", cleaned, re.IGNORECASE):
        return False

    return True


# ── Type inference ──────────────────────────────────────────────────


def infer_column_type(values: list) -> str:
    """Infer the BigQuery type for a column by scanning all non-empty values (D6).

    Returns one of: INT64, FLOAT64, BOOL, DATE, TIMESTAMP, STRING.
    Loadability guard: after picking a candidate, verify every non-empty value
    passes convert_value; on failure demote to STRING.
    """
    non_empty = [v for v in values if v is not None]
    if not non_empty:
        return "STRING"

    candidate = _infer_column_type_raw(non_empty)

    # Loadability guard: if any value fails conversion, demote to STRING
    if candidate != "STRING":
        for v in non_empty:
            try:
                convert_value(v, candidate)
            except ConversionError:
                return "STRING"

    return candidate


def _infer_column_type_raw(values: list) -> str:
    """Raw inference without the loadability guard."""
    # Check if all values are of the same native type
    types = set()
    for v in values:
        if isinstance(v, bool):
            types.add("bool")
        elif isinstance(v, int):
            types.add("int")
        elif isinstance(v, float):
            types.add("float")
        elif isinstance(v, date) and not isinstance(v, datetime):
            types.add("date")
        elif isinstance(v, datetime):
            types.add("datetime")
        elif isinstance(v, str):
            types.add("str")
        else:
            types.add("other")

    # Single native type
    if len(types) == 1:
        t = types.pop()
        if t == "bool":
            return "BOOL"
        if t == "int":
            return "INT64"
        if t == "float":
            return "FLOAT64"
        if t == "date":
            return "DATE"
        if t == "datetime":
            return "TIMESTAMP"
        if t == "str":
            # Use bigquery.py's string inference for consistency
            return _infer_str_column_type(values)
        return "STRING"

    # Mixed numeric types: int+float → FLOAT64
    if types == {"int", "float"}:
        return "FLOAT64"

    # Any other mix → STRING
    return "STRING"


def _infer_str_column_type(values: list[str]) -> str:
    """Infer type from a list of strings.

    Uses bigquery.py's regexes for DATE/TIMESTAMP/INTEGER/FLOAT,
    but adds case-insensitive bool detection (bigquery.py's _infer_string_type
    is case-sensitive and returns STRING for 'true'/'false').
    """
    inferred = "STRING"
    for v in values:
        stripped = v.strip()
        if not stripped:
            continue

        # Case-insensitive bool check (supplements bigquery.py)
        if stripped.lower() in ("true", "false"):
            mapped = "BOOL"
        else:
            v_type = _infer_string_type(v)
            mapped = _BQ_TYPE_REVERSE.get(v_type, "STRING")

        inferred = _promote_type(inferred, mapped)
        if inferred == "STRING":
            break
    return inferred


def _promote_type(a: str, b: str) -> str:
    """Return the more specific (less permissive) of two internal BigQuery types.

    For inference: if all values are integers, we want INT64, not STRING.
    The type with the LOWER rank is more specific.
    """
    # Map internal types to bigquery.py ranks for comparison
    a_bq = _BQ_TYPE_MAP.get(a, a)
    b_bq = _BQ_TYPE_MAP.get(b, b)
    a_rank = _BQ_TYPE_RANK.get(a_bq, 9)
    b_rank = _BQ_TYPE_RANK.get(b_bq, 9)
    # Lower rank = more specific; return the more specific type
    result_bq = a_bq if a_rank <= b_rank else b_bq
    return _BQ_TYPE_REVERSE.get(result_bq, result_bq)


# ── Value conversion ────────────────────────────────────────────────


def convert_value(value, target: str, *, column_name: str = "", row_number: int = 0):
    """Convert a native Python value to a JSON-serializable form for the target BQ type (D7).

    Returns None for None/empty input for every target.
    Raises ConversionError on first incompatible value.
    """
    if value is None:
        return None

    # bool MUST be checked before int (isinstance(True, int) is True in Python)
    if isinstance(value, bool):
        if target == "BOOL":
            return value
        if target == "STRING":
            return str(value)
        raise ConversionError(
            column_name, row_number, value, target,
            f"bool value {value!r} is not compatible with {target}"
        )

    if isinstance(value, int):
        if target == "INT64":
            return value
        if target == "FLOAT64":
            return float(value)
        if target == "STRING":
            return str(value)
        raise ConversionError(
            column_name, row_number, value, target,
            f"int value {value!r} is not compatible with {target}"
        )

    if isinstance(value, float):
        if not math.isfinite(value):
            raise ConversionError(
                column_name, row_number, value, target,
                f"non-finite float {value!r} is not compatible with {target}"
            )
        if target == "INT64":
            if value.is_integer():
                return int(value)
            raise ConversionError(
                column_name, row_number, value, target,
                f"float {value!r} is not an integer"
            )
        if target == "FLOAT64":
            return value
        if target == "STRING":
            return str(value)
        raise ConversionError(
            column_name, row_number, value, target,
            f"float value {value!r} is not compatible with {target}"
        )

    if isinstance(value, date) and not isinstance(value, datetime):
        if target == "DATE":
            return value.isoformat()
        if target == "STRING":
            return value.isoformat()
        raise ConversionError(
            column_name, row_number, value, target,
            f"date value {value!r} is not compatible with {target}"
        )

    if isinstance(value, datetime):
        if target == "TIMESTAMP":
            # Naive datetimes treated as UTC
            if value.tzinfo is None:
                return value.replace(tzinfo=timezone.utc).isoformat()
            return value.isoformat()
        if target == "STRING":
            return value.isoformat()
        raise ConversionError(
            column_name, row_number, value, target,
            f"datetime value {value!r} is not compatible with {target}"
        )

    if isinstance(value, str):
        return _convert_str_value(value, target, column_name, row_number)

    # Fallback for any other type
    if target == "STRING":
        return str(value)
    raise ConversionError(
        column_name, row_number, value, target,
        f"value {value!r} of type {type(value).__name__} is not compatible with {target}"
    )


def _convert_str_value(value: str, target: str, column_name: str, row_number: int):
    """Convert a string value to the target type."""
    stripped = value.strip()

    if target == "STRING":
        return value

    if target == "BOOL":
        if stripped.lower() in ("true", "false"):
            return stripped.lower() == "true"
        raise ConversionError(
            column_name, row_number, value, target,
            f"string {value!r} is not a boolean"
        )

    if target == "INT64":
        if re.match(r"^[+-]?\d+$", stripped):
            return int(stripped)
        raise ConversionError(
            column_name, row_number, value, target,
            f"string {value!r} is not an integer"
        )

    if target == "FLOAT64":
        # Match decimal or scientific notation
        if re.match(r"^[+-]?\d+\.\d+([eE][+-]?\d+)?$", stripped) or \
           re.match(r"^[+-]?\d+[eE][+-]?\d+$", stripped):
            f = float(stripped)
            if math.isfinite(f):
                return f
            raise ConversionError(
                column_name, row_number, value, target,
                f"float {value!r} is not finite"
            )
        raise ConversionError(
            column_name, row_number, value, target,
            f"string {value!r} is not a float"
        )

    if target == "DATE":
        try:
            d = date.fromisoformat(stripped)
            return d.isoformat()
        except ValueError:
            raise ConversionError(
                column_name, row_number, value, target,
                f"string {value!r} is not a valid DATE (YYYY-MM-DD)"
            )

    if target == "TIMESTAMP":
        try:
            # Accept Z suffix by replacing with +00:00
            ts = stripped.replace("Z", "+00:00")
            dt = datetime.fromisoformat(ts)
            # Naive datetimes treated as UTC
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.isoformat()
        except ValueError:
            raise ConversionError(
                column_name, row_number, value, target,
                f"string {value!r} is not a valid TIMESTAMP"
            )

    raise ConversionError(
        column_name, row_number, value, target,
        f"string {value!r} is not compatible with {target}"
    )
