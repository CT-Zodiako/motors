"""API tests for POST /bigquery/upload-file/inspect and /preview (PR2).

Fully offline: no BigQuery client is involved in these endpoints.
Multipart posts via the shared `client` fixture (conftest).
"""
import io

import pytest


# ── fixtures ─────────────────────────────────────────────────────────

def _csv_bytes(text: str) -> bytes:
    return text.encode("utf-8")


def _xlsx_bytes(sheets: dict[str, list[list]]) -> bytes:
    """Build an in-memory workbook: {sheet_name: [[row], [row], ...]}."""
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        if first:
            ws.title = name
            first = False
        for r, row in enumerate(rows, start=1):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _post(client, path, content, filename, source_type, extra_data=None):
    data = {"sourceType": source_type}
    if extra_data:
        data.update(extra_data)
    return client.post(
        path,
        files={"file": (filename, content, "application/octet-stream")},
        data=data,
    )


def _post_inspect(client, content, filename, source_type):
    return _post(client, "/bigquery/upload-file/inspect", content, filename, source_type)


def _post_preview(client, content, filename, source_type, sheet=None):
    extra = {"sheet": sheet} if sheet is not None else None
    return _post(client, "/bigquery/upload-file/preview", content, filename, source_type, extra)


# ── upload guards (D11) ───────────────────────────────────────────────

class TestUploadGuards:
    def test_extension_source_type_mismatch_400(self, client):
        r = _post_inspect(client, _csv_bytes("a,b\n1,2\n"), "datos.csv", "xlsx")
        assert r.status_code == 400
        assert "does not match" in r.json()["detail"]

    def test_invalid_source_type_400(self, client):
        r = _post_inspect(client, _csv_bytes("a,b\n1,2\n"), "datos.csv", "json")
        assert r.status_code == 400
        assert "sourceType" in r.json()["detail"]

    def test_xls_extension_415(self, client):
        r = _post_inspect(client, b"\xd0\xcf\x11\xe0whatever", "viejo.xls", "xlsx")
        assert r.status_code == 415
        assert ".xls" in r.json()["detail"]

    def test_ole2_magic_renamed_csv_415(self, client):
        # A real .xls (OLE2) renamed to .csv must still be rejected.
        content = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64
        r = _post_inspect(client, content, "fake.csv", "csv")
        assert r.status_code == 415
        assert ".xls" in r.json()["detail"]

    def test_unknown_extension_415(self, client):
        r = _post_inspect(client, b"%PDF-1.4 fake", "datos.pdf", "csv")
        assert r.status_code == 415
        assert ".xlsx, .csv" in r.json()["detail"]

    def test_missing_extension_415(self, client):
        r = _post_inspect(client, _csv_bytes("a,b\n1,2\n"), "sinextension", "csv")
        assert r.status_code == 415

    def test_oversize_file_413(self, client):
        content = b"x" * (20 * 1024 * 1024 + 1)
        r = _post_inspect(client, content, "grande.csv", "csv")
        assert r.status_code == 413
        assert "20 MB" in r.json()["detail"]

    def test_row_cap_413(self, client):
        content = _csv_bytes("a,b\n" + "1,2\n" * 100_001)
        r = _post_preview(client, content, "muchas.csv", "csv")
        assert r.status_code == 413
        assert "100000" in r.json()["detail"]


# ── inspect ───────────────────────────────────────────────────────────

class TestInspect:
    def test_inspect_xlsx_multi_sheet(self, client):
        content = _xlsx_bytes({
            "Hoja1": [["a"], [1]],
            "Datos": [["b"], [2]],
            "Resumen": [["c"], [3]],
        })
        r = _post_inspect(client, content, "ventas.xlsx", "xlsx")
        assert r.status_code == 200
        body = r.json()
        assert body["sourceType"] == "xlsx"
        assert body["fileName"] == "ventas.xlsx"
        assert body["sizeBytes"] == len(content)
        assert body["sheets"] == ["Hoja1", "Datos", "Resumen"]
        assert body["sheetCount"] == 3

    def test_inspect_csv_pseudo_sheet(self, client):
        content = _csv_bytes("a,b\n1,2\n")
        r = _post_inspect(client, content, "datos.csv", "csv")
        assert r.status_code == 200
        body = r.json()
        assert body["sourceType"] == "csv"
        assert body["sheets"] == ["CSV"]
        assert body["sheetCount"] == 1

    def test_inspect_ragged_csv_fails_fast_400(self, client):
        r = _post_inspect(client, _csv_bytes("a,b\n1,2,3\n"), "roto.csv", "csv")
        assert r.status_code == 400
        assert "row 1" in r.json()["detail"]

    def test_inspect_corrupt_xlsx_400(self, client):
        r = _post_inspect(client, b"PK\x03\x04 not really a zip", "roto.xlsx", "xlsx")
        assert r.status_code == 400


# ── preview ───────────────────────────────────────────────────────────

class TestPreview:
    def test_preview_csv_contract(self, client):
        content = _csv_bytes("Nombre,edad,fecha\nAna,30,2024-01-15\nLuis,41,2024-02-20\n")
        r = _post_preview(client, content, "datos.csv", "csv")
        assert r.status_code == 200
        body = r.json()
        assert body["sheet"] == "CSV"
        assert body["totalRows"] == 2
        assert body["columns"] == [
            {"source": "Nombre", "name": "Nombre", "type": "STRING", "included": True},
            {"source": "edad", "name": "edad", "type": "INT64", "included": True},
            {"source": "fecha", "name": "fecha", "type": "DATE", "included": True},
        ]
        assert body["sample"] == [["Ana", "30", "2024-01-15"], ["Luis", "41", "2024-02-20"]]

    def test_preview_sanitizes_names(self, client):
        content = _csv_bytes("Fecha de Venta,precio unitario\n2024-01-15,10.5\n")
        r = _post_preview(client, content, "datos.csv", "csv")
        names = [c["name"] for c in r.json()["columns"]]
        sources = [c["source"] for c in r.json()["columns"]]
        assert names == ["Fecha_de_Venta", "precio_unitario"]
        assert sources == ["Fecha de Venta", "precio unitario"]

    def test_preview_sample_capped_at_100(self, client):
        rows = "".join(f"name{i},{i}\n" for i in range(150))
        content = _csv_bytes("nombre,num\n" + rows)
        r = _post_preview(client, content, "datos.csv", "csv")
        body = r.json()
        assert len(body["sample"]) == 100
        assert body["sample"][0] == ["name0", "0"]
        assert body["totalRows"] == 150

    def test_preview_header_only_csv(self, client):
        r = _post_preview(client, _csv_bytes("a,b\n"), "datos.csv", "csv")
        assert r.status_code == 200
        body = r.json()
        assert body["totalRows"] == 0
        assert body["sample"] == []
        assert len(body["columns"]) == 2

    def test_preview_empty_rows_not_counted(self, client):
        r = _post_preview(client, _csv_bytes("a,b\n1,2\n\n\n3,4\n"), "datos.csv", "csv")
        assert r.json()["totalRows"] == 2

    def test_preview_ragged_csv_400_with_row_number(self, client):
        r = _post_preview(client, _csv_bytes("a,b,c\n1,2,3\n4,5\n"), "roto.csv", "csv")
        assert r.status_code == 400
        assert "row 2" in r.json()["detail"]

    def test_preview_empty_csv_400(self, client):
        r = _post_preview(client, b"", "vacio.csv", "csv")
        assert r.status_code == 400

    def test_preview_xlsx_defaults_first_sheet(self, client):
        content = _xlsx_bytes({
            "Primera": [["nombre"], ["Ana"]],
            "Segunda": [["otro"], ["x"]],
        })
        r = _post_preview(client, content, "datos.xlsx", "xlsx")
        assert r.status_code == 200
        body = r.json()
        assert body["sheet"] == "Primera"
        assert body["columns"][0]["source"] == "nombre"
        assert body["totalRows"] == 1

    def test_preview_xlsx_explicit_sheet(self, client):
        content = _xlsx_bytes({
            "Primera": [["nombre"], ["Ana"]],
            "Segunda": [["precio"], [10.5]],
        })
        r = _post_preview(client, content, "datos.xlsx", "xlsx", sheet="Segunda")
        assert r.status_code == 200
        body = r.json()
        assert body["sheet"] == "Segunda"
        assert body["columns"][0]["type"] == "FLOAT64"

    def test_preview_xlsx_unknown_sheet_400(self, client):
        content = _xlsx_bytes({"Hoja1": [["a"], [1]]})
        r = _post_preview(client, content, "datos.xlsx", "xlsx", sheet="NoExiste")
        assert r.status_code == 400
        assert "Hoja1" in r.json()["detail"]

    def test_preview_xlsx_long_row_non_empty_extras_400(self, client):
        # Data row longer than the header with a non-empty extra cell → 400.
        # (Trailing all-empty extras are tolerated artifacts — covered by PR1 unit tests.)
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "H"
        ws.cell(row=1, column=1, value="a")
        ws.cell(row=1, column=2, value="b")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value=2)
        ws.cell(row=2, column=3, value="sobra")
        bio = io.BytesIO()
        wb.save(bio)
        r = _post_preview(client, bio.getvalue(), "datos.xlsx", "xlsx")
        assert r.status_code == 400
        assert "expected 2" in r.json()["detail"]

    def test_preview_xlsx_dates_serialized_iso(self, client):
        import datetime as dt

        content = _xlsx_bytes({"H": [["fecha"], [dt.datetime(2024, 1, 15, 10, 30)]]})
        r = _post_preview(client, content, "datos.xlsx", "xlsx")
        body = r.json()
        assert body["columns"][0]["type"] == "TIMESTAMP"
        assert body["sample"][0][0] == "2024-01-15T10:30:00"
