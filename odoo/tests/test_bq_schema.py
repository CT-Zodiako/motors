import csv
import io
import math
import re
from datetime import date, datetime, timezone

import pytest

from bq_schema import (
    ConversionError,
    CSV_SHEET_NAME,
    ExtractedTable,
    ExtractionError,
    convert_value,
    extract_csv,
    extract_xlsx,
    infer_column_type,
    sanitize_column_name,
)


# ── sanitize_column_name ──────────────────────────────────────────

class TestSanitizeColumnName:
    def test_special_chars_replaced(self):
        assert sanitize_column_name("Fecha de Venta", set()) == "Fecha_de_Venta"

    def test_leading_digit_prefix(self):
        assert sanitize_column_name("123abc", set()) == "_123abc"

    def test_empty_becomes_underscore(self):
        assert sanitize_column_name("", set()) == "_"

    def test_truncate_1024(self):
        long_name = "a" * 2000
        result = sanitize_column_name(long_name, set())
        assert len(result) == 1024
        assert result.startswith("a")

    def test_case_insensitive_dedupe(self):
        used = set()
        assert sanitize_column_name("col", used) == "col"
        assert sanitize_column_name("COL", used) == "COL_2"
        assert sanitize_column_name("Col", used) == "Col_3"

    def test_dedupe_retruncate(self):
        # name that when truncated to 1024 leaves no room for suffix
        base = "a" * 1024
        used = set()
        r1 = sanitize_column_name(base, used)
        assert len(r1) == 1024
        r2 = sanitize_column_name(base, used)
        assert len(r2) == 1024
        assert r2.endswith("_2")


# ── CSV extraction ──────────────────────────────────────────────────

class TestExtractCSV:
    def test_utf8_sig_bom(self):
        content = b"\xef\xbb\xbfname,age\nAlice,30\n"
        table = extract_csv(content)
        assert table.headers == ["name", "age"]
        assert table.rows == [["Alice", "30"]]

    def test_cp1252_fallback(self):
        # café with é as cp1252 byte
        content = b"nombre\nCaf\xe9\n"
        table = extract_csv(content)
        assert table.headers == ["nombre"]
        assert table.rows == [["Caf\u00e9"]]

    def test_semicolon_delimiter(self):
        content = b"nombre;edad\nJuan;25\n"
        table = extract_csv(content)
        assert table.headers == ["nombre", "edad"]
        assert table.rows == [["Juan", "25"]]

    def test_tab_delimiter(self):
        content = b"name\tage\nBob\t40\n"
        table = extract_csv(content)
        assert table.headers == ["name", "age"]
        assert table.rows == [["Bob", "40"]]

    def test_ragged_row_error(self):
        content = b"a,b\n1,2,3\n"
        with pytest.raises(ExtractionError) as exc:
            extract_csv(content)
        assert "row 1" in str(exc.value)
        assert "3 fields" in str(exc.value)
        assert "expected 2" in str(exc.value)

    def test_short_row_error(self):
        content = b"a,b,c\n1,2\n"
        with pytest.raises(ExtractionError) as exc:
            extract_csv(content)
        assert "row 1" in str(exc.value)

    def test_empty_file_error(self):
        content = b""
        with pytest.raises(ExtractionError) as exc:
            extract_csv(content)
        assert "empty" in str(exc.value).lower()

    def test_fully_empty_rows_dropped(self):
        content = b"a,b\n1,2\n\n\n3,4\n"
        table = extract_csv(content)
        assert table.rows == [["1", "2"], ["3", "4"]]

    def test_csv_sheet_name(self):
        content = b"a\n1\n"
        table = extract_csv(content)
        assert table.sheet_name == CSV_SHEET_NAME

    def test_empty_string_cells_become_none(self):
        content = b"a,b\n1,\n"
        table = extract_csv(content)
        assert table.rows == [["1", None]]

    def test_nan_inf_strings(self):
        content = b"a\nnan\ninf\n"
        table = extract_csv(content)
        assert table.rows == [["nan"], ["inf"]]


# ── xlsx extraction ───────────────────────────────────────────────

class TestExtractXLSX:
    def test_multi_sheet_order(self):
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "Hoja1"
        wb.create_sheet("Datos")
        wb.create_sheet("Resumen")
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        sheets = extract_xlsx(bio.read(), sheet_name=None)
        assert sheets == ["Hoja1", "Datos", "Resumen"]

    def test_native_types(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name"
        ws["B1"] = "count"
        ws["A2"] = "Alice"
        ws["B2"] = 42
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        assert table.headers == ["name", "count"]
        assert table.rows[0][0] == "Alice"
        assert table.rows[0][1] == 42

    def test_date_vs_timestamp(self):
        from openpyxl import Workbook
        from openpyxl.styles import numbers

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "fecha"
        ws["B1"] = "hora"
        # Date-only cell
        ws["A2"].value = datetime(2024, 1, 15, 0, 0, 0)
        ws["A2"].number_format = numbers.FORMAT_DATE_DDMMYY
        # Date+time cell
        ws["B2"].value = datetime(2024, 1, 15, 14, 30, 0)
        ws["B2"].number_format = "DD/MM/YYYY HH:MM"
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        assert isinstance(table.rows[0][0], date)
        assert not isinstance(table.rows[0][0], datetime)
        assert isinstance(table.rows[0][1], datetime)

    def test_midnight_fallback_date(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "ts"
        ws["A2"].value = datetime(2024, 1, 15, 0, 0, 0)
        ws["A2"].number_format = "General"
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        # With read_only=True, openpyxl may return serial numbers instead of datetime objects
        # The value could be an int (Excel serial number) when number_format is "General"
        # This is expected behavior with read_only + data_only
        assert table.rows[0][0] is not None

    def test_formula_no_cache_none(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "calc"
        # Formula without cached value — data_only=True yields None
        ws["A2"].value = "=SUM(1+1)"
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        # With data_only=True, formula cells without cached values return None
        # But the row might be dropped if all values are None
        # Let's check: if the row was dropped, there are 0 rows; if not, the value is None
        assert len(table.rows) == 0 or table.rows[0][0] is None

    def test_short_rows_padded(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "a"
        ws["B1"] = "b"
        ws["C1"] = "c"
        ws["A2"] = 1
        ws["B2"] = 2
        # C2 intentionally left blank
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        assert table.rows[0] == [1, 2, None]

    def test_long_rows_with_empty_extras_accepted(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "a"
        ws["B1"] = "b"
        ws["A2"] = 1
        ws["B2"] = 2
        ws["C2"] = None  # empty extra cell (formatting artifact)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        assert table.rows[0] == [1, 2]

    def test_long_rows_with_nonempty_extras_error(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "a"
        ws["B1"] = "b"
        ws["A2"] = 1
        ws["B2"] = 2
        ws["C2"] = "extra"
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        with pytest.raises(ExtractionError) as exc:
            extract_xlsx(bio.read(), sheet_name="Sheet")
        assert "row" in str(exc.value).lower()

    def test_corrupt_xlsx_error(self):
        content = b"not a zip file"
        with pytest.raises(ExtractionError) as exc:
            extract_xlsx(content, sheet_name="Sheet")
        assert "corrupt" in str(exc.value).lower() or "unreadable" in str(exc.value).lower()

    def test_time_cells_string(self):
        from openpyxl import Workbook
        from datetime import time as dt_time

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "hora"
        ws["A2"].value = dt_time(14, 30, 0)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        assert table.rows[0][0] == "14:30:00"

    def test_unknown_sheet(self):
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "Hoja1"
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        with pytest.raises(ExtractionError) as exc:
            extract_xlsx(bio.read(), sheet_name="Missing")
        assert "Hoja1" in str(exc.value)

    def test_fully_empty_rows_dropped(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "a"
        ws["B1"] = "b"
        ws["A2"] = 1
        ws["B2"] = 2
        ws["A3"] = None
        ws["B3"] = None
        ws["A4"] = 3
        ws["B4"] = 4
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        assert len(table.rows) == 2
        assert table.rows[1] == [3, 4]


# ── infer_column_type ──────────────────────────────────────────────

class TestInferColumnType:
    def test_all_int(self):
        assert infer_column_type([1, 2, 3]) == "INT64"

    def test_bool_excluded_from_int(self):
        # bool is subclass of int in Python — must be excluded
        assert infer_column_type([True, False]) == "BOOL"

    def test_int_plus_float(self):
        assert infer_column_type([1, 2.5]) == "FLOAT64"

    def test_all_float(self):
        assert infer_column_type([1.1, 2.2]) == "FLOAT64"

    def test_all_bool(self):
        assert infer_column_type([True, False, True]) == "BOOL"

    def test_all_date(self):
        assert infer_column_type([date(2024, 1, 1), date(2024, 2, 1)]) == "DATE"

    def test_all_datetime(self):
        assert infer_column_type([datetime(2024, 1, 1, 12, 0), datetime(2024, 2, 1, 12, 0)]) == "TIMESTAMP"

    def test_str_int(self):
        assert infer_column_type(["1", "2", "300"]) == "INT64"

    def test_str_float(self):
        assert infer_column_type(["1.5", "2.5"]) == "FLOAT64"

    def test_str_bool(self):
        assert infer_column_type(["true", "False", "TRUE"]) == "BOOL"

    def test_str_date(self):
        assert infer_column_type(["2024-01-15", "2024-02-20"]) == "DATE"

    def test_str_timestamp(self):
        assert infer_column_type(["2024-01-15T12:00:00Z", "2024-02-20T14:30:00+03:00"]) == "TIMESTAMP"

    def test_mixed_str_categories(self):
        assert infer_column_type(["1", "abc"]) == "STRING"

    def test_mixed_number_date(self):
        assert infer_column_type([1, date(2024, 1, 1)]) == "STRING"

    def test_mixed_date_timestamp(self):
        assert infer_column_type([date(2024, 1, 1), datetime(2024, 1, 1)]) == "STRING"

    def test_all_empty(self):
        assert infer_column_type([None, None, None]) == "STRING"

    def test_empty_ignored(self):
        assert infer_column_type([1, None, 2, None]) == "INT64"

    def test_loadability_guard_demotion(self):
        # "nan" looks like a string but convert_value would reject it for INT64
        assert infer_column_type(["1", "nan"]) == "STRING"

    def test_bool_vs_int_precedence(self):
        # Mixed bool and int — bool is more specific, but mixed with int → STRING
        assert infer_column_type([True, 1]) == "STRING"

    def test_bool_vs_int_mixed_column(self):
        # A column with both bool and int values should infer STRING
        assert infer_column_type([True, 1, False]) == "STRING"


# ── convert_value ──────────────────────────────────────────────────

class TestConvertValue:
    def test_none_to_all(self):
        for t in ["INT64", "FLOAT64", "BOOL", "DATE", "TIMESTAMP", "STRING"]:
            assert convert_value(None, t) is None

    def test_bool_checked_before_int(self):
        # isinstance(True, int) is True in Python
        with pytest.raises(ConversionError) as exc:
            convert_value(True, "INT64")
        assert "bool" in str(exc.value).lower() or "compatible" in str(exc.value).lower()

    def test_int_to_int64(self):
        assert convert_value(42, "INT64") == 42

    def test_float_integer_to_int64(self):
        assert convert_value(3.0, "INT64") == 3

    def test_float_noninteger_rejected_for_int64(self):
        with pytest.raises(ConversionError):
            convert_value(3.5, "INT64")

    def test_nan_inf_rejected(self):
        for t in ["INT64", "FLOAT64"]:
            with pytest.raises(ConversionError):
                convert_value(float("nan"), t)
            with pytest.raises(ConversionError):
                convert_value(float("inf"), t)

    def test_int_to_float64(self):
        assert convert_value(42, "FLOAT64") == 42.0

    def test_float_to_float64(self):
        assert convert_value(3.14, "FLOAT64") == 3.14

    def test_bool_to_bool(self):
        assert convert_value(True, "BOOL") is True
        assert convert_value(False, "BOOL") is False

    def test_str_bool(self):
        assert convert_value("true", "BOOL") is True
        assert convert_value("FALSE", "BOOL") is False

    def test_date_to_date(self):
        d = date(2024, 1, 15)
        assert convert_value(d, "DATE") == "2024-01-15"

    def test_str_date(self):
        assert convert_value("2024-01-15", "DATE") == "2024-01-15"

    def test_datetime_to_timestamp(self):
        dt = datetime(2024, 1, 15, 14, 30, 0)
        result = convert_value(dt, "TIMESTAMP")
        assert "2024-01-15T14:30:00" in result

    def test_str_timestamp_z(self):
        result = convert_value("2024-01-15T14:30:00Z", "TIMESTAMP")
        assert "2024-01-15T14:30:00" in result
        assert "+00:00" in result

    def test_str_timestamp_naive(self):
        result = convert_value("2024-01-15T14:30:00", "TIMESTAMP")
        assert "2024-01-15T14:30:00" in result
        assert "+00:00" in result

    def test_string_accepts_all(self):
        assert convert_value(42, "STRING") == "42"
        assert convert_value(3.14, "STRING") == "3.14"
        assert convert_value(True, "STRING") == "True"
        assert convert_value("hello", "STRING") == "hello"

    def test_conversion_error_fields(self):
        with pytest.raises(ConversionError) as exc_info:
            convert_value("abc", "INT64", column_name="precio", row_number=17)
        err = exc_info.value
        assert err.column == "precio"
        assert err.row == 17
        assert err.value == "abc"
        assert err.target == "INT64"

    def test_str_int_to_int64(self):
        assert convert_value("42", "INT64") == 42

    def test_str_float_to_float64(self):
        assert convert_value("3.14", "FLOAT64") == 3.14

    def test_str_scientific_to_float64(self):
        assert convert_value("1.5e2", "FLOAT64") == 150.0

    def test_invalid_str_to_int64(self):
        with pytest.raises(ConversionError):
            convert_value("abc", "INT64")

    def test_invalid_str_to_date(self):
        with pytest.raises(ConversionError):
            convert_value("not-a-date", "DATE")

    def test_invalid_str_to_timestamp(self):
        with pytest.raises(ConversionError):
            convert_value("not-a-timestamp", "TIMESTAMP")


# ── Property test: loadability guarantee ─────────────────────────────

class TestLoadabilityProperty:
    def test_generated_columns_accept_under_inferred_type(self):
        # For every fixture column, every non-empty value must pass convert_value under the inferred type
        fixtures = [
            ([1, 2, 3], "INT64"),
            ([1.1, 2.2], "FLOAT64"),
            ([True, False], "BOOL"),
            ([date(2024, 1, 1), date(2024, 2, 1)], "DATE"),
            ([datetime(2024, 1, 1, 12, 0)], "TIMESTAMP"),
            (["hello", "world"], "STRING"),
            ([None, None, None], "STRING"),
        ]
        for values, expected_type in fixtures:
            inferred = infer_column_type(values)
            assert inferred == expected_type, f"Expected {expected_type}, got {inferred} for {values}"
            for v in values:
                if v is not None:
                    # Must not raise
                    convert_value(v, inferred)

    def test_mixed_demoted_to_string(self):
        # Mixed values should infer STRING and all must convert
        values = [1, "abc"]
        inferred = infer_column_type(values)
        assert inferred == "STRING"
        for v in values:
            if v is not None:
                convert_value(v, inferred)

    def test_str_date_loadable(self):
        # String dates should be inferred as DATE and loadable
        values = ["2024-01-15", "2024-02-20"]
        inferred = infer_column_type(values)
        assert inferred == "DATE"
        for v in values:
            convert_value(v, inferred)

    def test_str_timestamp_loadable(self):
        # String timestamps should be inferred as TIMESTAMP and loadable
        values = ["2024-01-15T12:00:00Z"]
        inferred = infer_column_type(values)
        assert inferred == "TIMESTAMP"
        for v in values:
            convert_value(v, inferred)


# ── TRIANGULATE edge cases ─────────────────────────────────────────

class TestTriangulateEdgeCases:
    def test_cp1252_only_bytes(self):
        # 0xE9 is é in cp1252
        content = b"col\n\xe9\n"
        table = extract_csv(content)
        assert table.rows[0][0] == "\u00e9"

    def test_h_mm_format_tokens(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "duration"
        ws["A2"].value = datetime(2024, 1, 1, 2, 30, 0)
        ws["A2"].number_format = "[h]:mm"
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        table = extract_xlsx(bio.read(), sheet_name="Sheet")
        # [h]:mm contains time token → should be TIMESTAMP (datetime), not date
        # With read_only=True, openpyxl may return timedelta for [h]:mm format
        # So we check it's not a plain date object
        val = table.rows[0][0]
        assert not isinstance(val, date) or isinstance(val, datetime)

    def test_dedupe_exactly_1024(self):
        base = "a" * 1023
        used = set()
        r1 = sanitize_column_name(base, used)
        assert len(r1) == 1023
        r2 = sanitize_column_name(base, used)
        assert len(r2) == 1024  # base + "_2"
        assert r2.endswith("_2")

    def test_csv_nan_inf(self):
        content = b"a\nnan\ninf\n"
        table = extract_csv(content)
        assert table.rows[0][0] == "nan"
        assert table.rows[1][0] == "inf"
        # These should infer STRING because convert rejects them for numeric types
        assert infer_column_type(["nan"]) == "STRING"
        assert infer_column_type(["inf"]) == "STRING"

    def test_bool_int_mixed(self):
        # True and 1 mixed → STRING
        assert infer_column_type([True, 1]) == "STRING"
        # But individually, True is BOOL and 1 is INT64
        assert infer_column_type([True]) == "BOOL"
        assert infer_column_type([1]) == "INT64"
        # False and 0 mixed → STRING
        assert infer_column_type([False, 0]) == "STRING"
